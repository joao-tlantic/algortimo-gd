import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from ortools.sat.python import cp_model
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import logging
from typing import Dict, Any, List, Tuple, Optional, Callable
from base_data_project.log_config import get_logger
from src.config import PROJECT_NAME
from src.config import ROOT_DIR
import os

# Set up logger
logger = get_logger(PROJECT_NAME)

#----------------------------------------SOLVER-----------------------------------------------------------
def solve(
    model: cp_model.CpModel, 
    days_of_year: List[int], 
    workers: List[int], 
    special_days: List[int], 
    shift: Dict[Tuple[int, int, str], cp_model.IntVar], 
    shifts: List[str],
    max_time_seconds: int = 240,
    enumerate_all_solutions: bool = False,
    use_phase_saving: bool = True,
    log_search_progress: bool = True,
    log_callback: Optional[Callable] = None,
    output_filename: str = os.path.join(ROOT_DIR, 'data', 'output', 'working_schedule.xlsx')
) -> pd.DataFrame:
    """
    Enhanced solver function with comprehensive logging and configurable parameters.
    
    Args:
        model: The CP-SAT model to solve
        days_of_year: List of days to schedule
        workers: List of worker IDs
        special_days: List of special days (holidays, sundays)
        shift: Dictionary mapping (worker, day, shift) to decision variables
        shifts: List of available shift types
        max_time_seconds: Maximum solving time in seconds (default: 240)
        enumerate_all_solutions: Whether to enumerate all solutions (default: False)
        use_phase_saving: Whether to use phase saving (default: True)
        log_search_progress: Whether to log search progress (default: True)
        log_callback: Custom callback for logging (default: None, uses print)
        output_filename: Name of the output Excel file (default: 'worker_schedule.xlsx')
    
    Returns:
        DataFrame containing the worker schedule
        
    Raises:
        ValueError: If input parameters are invalid
        RuntimeError: If solver fails to find a solution
    """
    try:
        logger.info("Starting enhanced solver with comprehensive logging")
        
        # =================================================================
        # 1. VALIDATE INPUT PARAMETERS
        # =================================================================
        logger.info("Validating input parameters")
        
        if not isinstance(model, cp_model.CpModel):
            raise ValueError("model must be a CP-SAT CpModel instance")
        
        if not days_of_year or not isinstance(days_of_year, list):
            raise ValueError("days_of_year must be a non-empty list")
        
        if not workers or not isinstance(workers, list):
            raise ValueError("workers must be a non-empty list")
        
        if not shifts or not isinstance(shifts, list):
            raise ValueError("shifts must be a non-empty list")
        
        if not isinstance(shift, dict):
            raise ValueError("shift must be a dictionary")
        
        if max_time_seconds <= 0:
            raise ValueError("max_time_seconds must be positive")
        
        logger.info(f"✅ Input validation passed:")
        logger.info(f"  - Days to schedule: {len(days_of_year)} days (from {min(days_of_year)} to {max(days_of_year)})")
        logger.info(f"  - Workers: {len(workers)} workers")
        logger.info(f"  - Special days: {len(special_days)} days")
        logger.info(f"  - Available shifts: {shifts}")
        logger.info(f"  - Decision variables: {len(shift)} variables")
        logger.info(f"  - Max solving time: {max_time_seconds} seconds")
        
        # =================================================================
        # 2. CONFIGURE AND CREATE SOLVER
        # =================================================================
        logger.info("Configuring CP-SAT solver")
        
        solver = cp_model.CpSolver()
        solver.parameters.max_time_in_seconds = max_time_seconds
        solver.parameters.enumerate_all_solutions = enumerate_all_solutions
        solver.parameters.use_phase_saving = use_phase_saving
        solver.parameters.log_search_progress = log_search_progress
        
        # Set up logging callback
        if log_callback is None:
            log_callback = lambda x: logger.info(f"Solver progress: {x}")
        solver.log_callback = log_callback
        
        logger.info(f"Solver configuration:")
        logger.info(f"  - Max time: {max_time_seconds} seconds")
        logger.info(f"  - Enumerate all solutions: {enumerate_all_solutions}")
        logger.info(f"  - Use phase saving: {use_phase_saving}")
        logger.info(f"  - Log search progress: {log_search_progress}")
        
        # =================================================================
        # 3. SOLVE THE PROBLEM
        # =================================================================
        logger.info("Starting optimization process...")
        start_time = datetime.now()
        
        status = solver.Solve(model)
        
        end_time = datetime.now()
        solve_duration = (end_time - start_time).total_seconds()
        
        logger.info(f"Optimization completed in {solve_duration:.2f} seconds")
        logger.info(f"Solver status: {solver.status_name(status)}")
        
        # Log solver statistics
        logger.info(f"Solver statistics:")
        logger.info(f"  - Objective value: {solver.ObjectiveValue() if status in [cp_model.OPTIMAL, cp_model.FEASIBLE] else 'N/A'}")
        logger.info(f"  - Best objective bound: {solver.BestObjectiveBound() if status in [cp_model.OPTIMAL, cp_model.FEASIBLE] else 'N/A'}")
        logger.info(f"  - Number of branches: {solver.NumBranches()}")
        logger.info(f"  - Number of conflicts: {solver.NumConflicts()}")
        logger.info(f"  - Wall time: {solver.WallTime():.2f} seconds")

        # =================================================================
        # 4. VALIDATE SOLUTION STATUS
        # =================================================================
        if status not in [cp_model.OPTIMAL, cp_model.FEASIBLE]:
            error_msg = f"Solver failed to find a solution. Status: {solver.status_name(status)}"
            logger.error(error_msg)
            
            if status == cp_model.INFEASIBLE:
                logger.error("Problem is infeasible - no solution exists with current constraints")
            elif status == cp_model.MODEL_INVALID:
                logger.error("Model is invalid - check constraint definitions")
            elif status == cp_model.UNKNOWN:
                logger.error("Solver timed out or encountered unknown status")
            
            raise RuntimeError(error_msg)
        
        logger.info(f"✅ Solution found! Status: {solver.status_name(status)}")
        
        # =================================================================
        # 5. PROCESS SOLUTION AND CREATE SCHEDULE
        # =================================================================
        logger.info("Processing solution and creating schedule")
        
        # Shift mapping for readability
        shift_mapping = {
            'M'     : 'M',  # Morning shift
            'T'     : 'T',  # Afternoon shift
            'F'     : 'F',  # Closed holiday
            'V'     : 'V',  # Empty Day
            'A'     : 'A',  # Missing shift
            'L'     : 'L',  # Free day
            'LQ'    : 'LQ', # Free days semester
            'LD'    : 'LD',
            'TC'    : 'TC',
        }
        
        logger.info(f"Shift mapping: {shift_mapping}")

        # Prepare the data for the DataFrame
        table_data = []  # List to store each worker's data as a row
        worker_stats = {}  # Dictionary to track L, LQ, LD counts for each worker
        
        logger.info(f"Processing schedule for {len(workers)} workers across {len(days_of_year)} days")
        # Prepare the data for the DataFrame
        table_data = []  # List to store each worker's data as a row
        worker_stats = {}  # Dictionary to track L, LQ, LD counts for each worker

        # Loop through each worker
        processed_workers = 0
        for w in workers:
            try:
                worker_row = [w]  # Start with the worker's name
                # Initialize counters for this worker
                l_count = 0
                lq_count = 0
                ld_count = 0
                tc_count = 0
                special_days_count = 0  # Counter for special days with M or T shifts
                unassigned_days = 0
                
                logger.debug(f"Processing worker {w}")

                days_of_year_sorted = sorted(days_of_year)
                for d in days_of_year_sorted:
                    assigned_shift = None  # To store the assigned shift (M, T, F)
                    
                    # Check if this is a special day
                    is_special_day = d in special_days
                    
                    # Check all shifts for this worker on this day
                    shifts_found = 0
                    for s in shifts:
                        try:
                            if (w, d, s) not in shift:
                                logger.warning(f"Missing decision variable for worker {w}, day {d}, shift {s}")
                                continue
                                
                            shift_value = solver.Value(shift[(w, d, s)])

                            if shift_value == 1:  # If the worker is assigned this shift
                                shifts_found += 1
                                assigned_shift = shift_mapping.get(s, s)  # Get the corresponding shift
                                
                                # Count L, LQ, LD shifts
                                if s == 'L':
                                    l_count += 1
                                elif s == 'LQ':
                                    lq_count += 1
                                elif s == 'LD':
                                    ld_count += 1
                                elif s == 'TC':
                                    tc_count += 1
                                    
                                # Count M or T shifts on special days
                                if is_special_day and (s == 'M' or s == 'T'):
                                    special_days_count += 1
                                    
                        except Exception as e:
                            logger.warning(f"Error processing shift for worker {w}, day {d}, shift {s}: {e}")
                            continue
                    
                    # Validate that exactly one shift is assigned per day
                    if shifts_found == 0:
                        unassigned_days += 1
                        assigned_shift = 'N'  # No shift assigned
                        logger.warning(f"No shift assigned to worker {w} on day {d}")
                    elif shifts_found > 1:
                        logger.warning(f"Multiple shifts assigned to worker {w} on day {d}")
                    
                    worker_row.append(assigned_shift or 'N')  # Add the shift or 'N' if none
                
                # Store the counts for this worker
                worker_stats[w] = {
                    'L': l_count, 
                    'LQ': lq_count, 
                    'LD': ld_count,
                    'TC': tc_count,
                    'Special_MT': special_days_count,
                    'Unassigned': unassigned_days
                }
                table_data.append(worker_row)  # Add the worker's data to the table
                processed_workers += 1
                
            except Exception as e:
                logger.error(f"Error processing worker {w}: {e}")
                # Add empty row for this worker to maintain structure
                worker_row = [w] + ['ERROR'] * len(days_of_year)
                table_data.append(worker_row)
                worker_stats[w] = {'L': 0, 'LQ': 0, 'LD': 0, 'TC': 0, 'Special_MT': 0, 'Unassigned': len(days_of_year)}

        logger.info(f"✅ Successfully processed {processed_workers}/{len(workers)} workers")

        # =================================================================
        # 6. CREATE AND VALIDATE DATAFRAME
        # =================================================================
        logger.info("Creating schedule DataFrame")
        
        try:
            # Create a DataFrame from the data
            column_headers = ['Worker'] + [f"Day {d}" for d in sorted(days_of_year)]
            schedule_df = pd.DataFrame(table_data, columns=column_headers)
            
            logger.info(f"DataFrame created: {schedule_df.shape}")
            logger.info(f"  - Rows (workers): {len(schedule_df)}")
            logger.info(f"  - Columns (days + worker): {len(schedule_df.columns)}")
            
        except Exception as e:
            logger.error(f"Error creating DataFrame: {e}")
            raise RuntimeError(f"Failed to create schedule DataFrame: {e}")

        # =================================================================
        # 7. LOG WORKER STATISTICS
        # =================================================================
        logger.info("Logging worker statistics")
        
        logger.info("Shift Counts by Worker:")
        logger.info("-" * 80)
        logger.info(f"{'Worker':<15} {'L':<5} {'LQ':<5} {'LD':<5} {'TC':<5} {'Total':<5} {'Special M/T':<10} {'Unassigned':<10}")
        logger.info("-" * 80)
        
        total_stats = {'L': 0, 'LQ': 0, 'LD': 0, 'TC': 0, 'Special_MT': 0, 'Unassigned': 0}
        
        for worker, counts in worker_stats.items():
            total_free = counts['L'] + counts['LQ'] + counts['LD']
            logger.info(f"{worker:<15} {counts['L']:<5} {counts['LQ']:<5} {counts['LD']:<5} {counts['TC']:<5} {total_free:<5} {counts['Special_MT']:<10} {counts['Unassigned']:<10}")
            
            # Accumulate totals
            for key in total_stats:
                total_stats[key] += counts[key]
        
        logger.info("-" * 80)
        logger.info(f"{'TOTAL':<15} {total_stats['L']:<5} {total_stats['LQ']:<5} {total_stats['LD']:<5} {total_stats['TC']:<5} {total_stats['L'] + total_stats['LQ'] + total_stats['LD']:<5} {total_stats['Special_MT']:<10} {total_stats['Unassigned']:<10}")
        logger.info("-" * 80)

        # =================================================================
        # 8. EXPORT TO EXCEL WITH ERROR HANDLING
        # =================================================================
        logger.info(f"Exporting schedule to Excel file: {output_filename}")
        
        try:
            # Export the DataFrame to an Excel file
            schedule_df.to_excel(output_filename, index=False)
            logger.info(f"✅ Initial Excel file created: {output_filename}")

            # Add formatting and summary sheet
            wb = load_workbook(output_filename)
            ws = wb.active  # Get the first worksheet

            # Create a fill pattern for special days (light blue background)
            special_day_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

            # Find column indices that correspond to days in special_days
            special_days_highlighted = 0
            for col_idx, column_name in enumerate(schedule_df.columns[1:], start=2):
                try:
                    # Extract day number from column name (format is "Day X")
                    day_num = int(column_name.split(' ')[1])
                    if day_num in special_days:
                        # Apply fill to the header cell (first row)
                        cell = ws.cell(row=1, column=col_idx)
                        cell.fill = special_day_fill
                        special_days_highlighted += 1
                        logger.debug(f"Highlighted Day {day_num} as special day in column {col_idx}")
                except (ValueError, IndexError) as e:
                    logger.warning(f"Could not parse day number from column: {column_name}, error: {e}")
                    continue

            logger.info(f"✅ Highlighted {special_days_highlighted} special days in Excel")

            # Save the changes
            wb.save(output_filename)
            
            # Add a summary sheet with statistics
            with pd.ExcelWriter(output_filename, engine='openpyxl', mode='a') as writer:
                # Create a DataFrame for the counts
                counts_df = pd.DataFrame([
                    {
                        'Worker': w, 
                        'L': stats['L'], 
                        'LQ': stats['LQ'], 
                        'LD': stats['LD'], 
                        'TC': stats['TC'],
                        'Total Free Days': stats['L'] + stats['LQ'] + stats['LD'],
                        'Special Days (M/T)': stats['Special_MT'],
                        'Unassigned Days': stats['Unassigned']
                    }
                    for w, stats in worker_stats.items()
                ])
                counts_df.to_excel(writer, sheet_name='Shift Summary', index=False)
                logger.info("✅ Summary sheet added to Excel file")

            logger.info(f"✅ Schedule successfully exported to '{output_filename}' with summary sheet")

        except Exception as e:
            logger.error(f"Error exporting to Excel: {e}")
            logger.warning("Continuing without Excel export...")

        # =================================================================
        # 9. FINAL VALIDATION AND RETURN
        # =================================================================
        logger.info("Performing final validation")
        
        # Check for any critical issues
        if total_stats['Unassigned'] > 0:
            logger.warning(f"Found {total_stats['Unassigned']} unassigned days across all workers")
        
        # Validate DataFrame integrity
        expected_rows = len(workers)
        expected_cols = len(days_of_year) + 1  # +1 for worker column
        
        if len(schedule_df) != expected_rows:
            logger.warning(f"DataFrame row count mismatch: expected {expected_rows}, got {len(schedule_df)}")
        
        if len(schedule_df.columns) != expected_cols:
            logger.warning(f"DataFrame column count mismatch: expected {expected_cols}, got {len(schedule_df.columns)}")

        logger.info("✅ Enhanced solver completed successfully")
        return schedule_df
        
    except Exception as e:
        logger.error(f"Error in enhanced solver: {e}", exc_info=True)
        raise