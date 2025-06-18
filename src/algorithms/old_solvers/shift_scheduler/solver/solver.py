import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from ortools.sat.python import cp_model
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

#----------------------------------------SOLVER-----------------------------------------------------------
def solve(model, days_of_year, workers, special_days, shift, shifts):
    # Create the solver
    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = 600
    solver.parameters.enumerate_all_solutions = False
    solver.parameters.use_phase_saving = True
    solver.parameters.log_search_progress = True
    solver.log_callback = print

    # Solve the problem
    status = solver.Solve(model)

    # Shift mapping for readability
    # Shift mapping for readability
    shift_mapping = {
        'M'     : 'M',  # Morning shift
        'T'     : 'T',  # Afternoon shift
        'F'     : 'F',  # Closed holiday
        'V'     : 'V',  # Empty Day
        'A'     : 'A',  # Missing shift
        'L'     : 'L',  # Free day
        'LQ'    : 'LQ', # Free days semester
        'LD'    : 'LD',
        'TC'    : 'TC',
    }

    # Check if a solution was found
    if status == cp_model.OPTIMAL or status == cp_model.FEASIBLE:
        # Prepare the data for the DataFrame
        table_data = []  # List to store each worker's data as a row
        worker_stats = {}  # Dictionary to track L, LQ, LD counts for each worker

        # Loop through each worker
        for w in workers:
            worker_row = [w]  # Start with the worker's name
            # Initialize counters for this worker
            l_count = 0
            lq_count = 0
            ld_count = 0
            tc_count = 0
            special_days_count = 0  # Counter for special days with M or T shifts
            

            days_of_year = sorted(days_of_year)
            for d in days_of_year:
                assigned_shift = None  # To store the assigned shift (M, T, F)
                
                # Check if this is a special day (you'll need to define what makes a day "special")
                is_special_day = d in special_days  # Assuming 'special_days' is defined elsewhere
                
                for s in shifts:
                    shift_value = solver.Value(shift[(w, d, s)])

                    if shift_value == 1:  # If the worker is assigned this shift
                        assigned_shift = shift_mapping.get(s, '')  # Get the corresponding shift
                        # Count L, LQ, LD shifts
                        if s == 'L':
                            l_count += 1
                        elif s == 'LQ':
                            lq_count += 1
                        elif s == 'LD':
                            ld_count += 1
                        elif s == 'TC':
                            tc_count += 1
                        # Count M or T shifts on special days
                        if is_special_day and (s == 'M' or s == 'T'):
                            special_days_count += 1
                        break  # Once we find the assigned shift, break out of the inner loop
                # If no shift is assigned, print a placeholder like 'N' for "None"
                worker_row.append(assigned_shift or 'N')  # Add the shift or 'N' if none
            
            # Store the counts for this worker
            worker_stats[w] = {
                'L': l_count, 
                'LQ': lq_count, 
                'LD': ld_count,
                'TC': tc_count,
                'Special_MT': special_days_count
            }
            table_data.append(worker_row)  # Add the worker's data to the table

        # Create a DataFrame from the data
        column_headers = ['Worker'] + [f"Day {d}" for d in days_of_year]  # Column headers: Worker + Days
        schedule_df = pd.DataFrame(table_data, columns=column_headers)

        # Print the L, LQ, LD counts and special days for each worker
        print("\nShift Counts by Worker:")
        print("-" * 65)
        print(f"{'Worker':<15} {'L':<5} {'LQ':<5} {'LD':<5} {'TC':<5} {'Total':<5} {'Special M/T':<10}")
        print("-" * 65)
        for worker, counts in worker_stats.items():
            total_free = counts['L'] + counts['LQ'] + counts['LD']
            print(f"{worker:<15} {counts['L']:<5} {counts['LQ']:<5} {counts['LD']:<5} {counts['TC']:<5} {total_free:<5} {counts['Special_MT']:<10}")
        print("-" * 65)

        # Debugging: print the DataFrame to check its contents
        print("\nFinal DataFrame:")
        print(schedule_df)

        # Export the DataFrame to an Excel file
        schedule_df.to_excel('worker_schedule.xlsx', index=False)

        # Now reopen the file to add formatting and the summary sheet

        wb = load_workbook('worker_schedule.xlsx')
        ws = wb.active  # Get the first worksheet

        # Create a fill pattern for special days (light blue background)
        special_day_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

        # Find column indices that correspond to days in special_days
        for col_idx, column_name in enumerate(schedule_df.columns[1:], start=2):  # Start from 2 because col 1 is worker names
            try:
                # Extract day number from column name (format is "Day X")
                day_num = int(column_name.split(' ')[1])  # Split by space and get the second part
                if day_num in special_days:
                    # Apply fill to the header cell (first row)
                    cell = ws.cell(row=1, column=col_idx)
                    cell.fill = special_day_fill
                    print(f"Highlighting Day {day_num} as special day in column {col_idx}")
            except (ValueError, IndexError):
                # Skip columns that don't have day numbers
                print(f"Could not parse day number from column: {column_name}")
                continue

        # Save the changes so far
        wb.save('worker_schedule.xlsx')
        
        # Also add a summary sheet with the L, LQ, LD counts and special days
        with pd.ExcelWriter('worker_schedule.xlsx', engine='openpyxl', mode='a') as writer:
            # Create a DataFrame for the counts
            counts_df = pd.DataFrame([
                {'Worker': w, 'L': stats['L'], 'LQ': stats['LQ'], 'LD': stats['LD'], 'TC': stats['TC'] ,
                'Total Free Days': stats['L'] + stats['LQ'] + stats['LD'],
                'Special Days (M/T)': stats['Special_MT']}
                for w, stats in worker_stats.items()
            ])
            counts_df.to_excel(writer, sheet_name='Shift Summary', index=False)

        print("Schedule has been exported to 'worker_schedule.xlsx' with summary sheet")

        # Print information about consecutive pairs
        print("\nConsecutive Pair Variable Values:")
        print("-" * 65)
        print(f"{'Worker':<10} {'Current Day':<12} {'Next Day':<10} {'Value':<6}")
        print("-" * 65)


    else:
        print("No solution found.")

    print(solver.status_name(status))


    return schedule_df